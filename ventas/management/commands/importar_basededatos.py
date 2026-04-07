from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
from catalogos.models import Categoria, Sucursal, MetodoPago, Cliente, Vendedor, Producto
from ventas.models import Venta, DetalleVenta

class Command(BaseCommand):
    help = 'Importa datos desde el archivo Excel DashBoard2021.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', type=str, required=True)

    def handle(self, *args, **options):
        archivo = options['archivo']
        wb = load_workbook(archivo, data_only=True)
        
        if 'BaseDe Datos' not in wb.sheetnames:
            self.stdout.write(self.style.ERROR("No existe la hoja 'BaseDe Datos'"))
            return
            
        ws = wb['BaseDe Datos']
        filas = list(ws.iter_rows(values_only=True))
        encabezados = filas[0]
        datos = filas[1:]
        
        ventas_temporales = {}
        
        for fila in datos:
            registro = dict(zip(encabezados, fila))
            
            categoria_nombre = str(registro.get('Categoria', '')).strip()
            sucursal_nombre = str(registro.get('Sucursal', '')).strip()
            ciudad = str(registro.get('Ciudad', '')).strip()
            estado = str(registro.get('Estado', '')).strip()
            pais = str(registro.get('Pais', '')).strip()
            metodo_pago_nombre = str(registro.get('Metodo Pago', '')).strip()
            cliente_nombre = str(registro.get('Cliente', '')).strip()
            vendedor_nombre = str(registro.get('Vendedor', '')).strip()
            codigo_producto = str(registro.get('CodigoProducto', '')).strip()
            nombre_producto = str(registro.get('Producto', '')).strip()
            folio = str(registro.get('Documento', '')).strip()
            
            fecha_valor = registro.get('Fecha')
            if isinstance(fecha_valor, datetime):
                fecha = fecha_valor.date()
            else:
                fecha = datetime.strptime(str(fecha_valor), '%Y-%m-%d').date()
                
            cantidad = Decimal(str(registro.get('Cantidad', 0) or 0))
            precio_unitario = Decimal(str(registro.get('PrecioUnitario', 0) or 0))
            importe = Decimal(str(registro.get('Importe', 0) or 0))
            subtotal = Decimal(str(registro.get('Subtotal', 0) or 0))
            impuesto = Decimal(str(registro.get('Impuesto', 0) or 0))
            total = Decimal(str(registro.get('Total', 0) or 0))
            
            categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
            
            sucursal, _ = Sucursal.objects.get_or_create(
                nombre=sucursal_nombre,
                defaults={'ciudad': ciudad, 'estado': estado, 'pais': pais}
            )
            
            metodo_pago, _ = MetodoPago.objects.get_or_create(nombre=metodo_pago_nombre)
            cliente, _ = Cliente.objects.get_or_create(nombre=cliente_nombre)
            
            vendedor, _ = Vendedor.objects.get_or_create(
                nombre=vendedor_nombre,
                defaults={'sucursal': sucursal}
            )
            
            producto, _ = Producto.objects.get_or_create(
                codigo=codigo_producto,
                defaults={
                    'nombre': nombre_producto,
                    'categoria': categoria,
                    'precio': precio_unitario
                }
            )
            
            if folio not in ventas_temporales:
                ventas_temporales[folio] = {
                    'fecha': fecha,
                    'cliente': cliente,
                    'sucursal': sucursal,
                    'vendedor': vendedor,
                    'metodo_pago': metodo_pago,
                    'subtotal': subtotal,
                    'impuesto': impuesto,
                    'total': total,
                    'detalles': []
                }
                
            ventas_temporales[folio]['detalles'].append({
                'producto': producto,
                'cantidad': cantidad,
                'precio_unitario': precio_unitario,
                'importe': importe
            })
            
        for folio, info in ventas_temporales.items():
            venta, creada = Venta.objects.get_or_create(
                folio=folio,
                defaults={
                    'fecha': info['fecha'],
                    'cliente': info['cliente'],
                    'sucursal': info['sucursal'],
                    'vendedor': info['vendedor'],
                    'metodo_pago': info['metodo_pago'],
                    'subtotal': info['subtotal'],
                    'impuesto': info['impuesto'],
                    'total': info['total'],
                }
            )
            
            if creada:
                for d in info['detalles']:
                    DetalleVenta.objects.create(
                        venta=venta,
                        producto=d['producto'],
                        cantidad=d['cantidad'],
                        precio_unitario=d['precio_unitario'],
                        importe=d['importe']
                    )
                    
        self.stdout.write(self.style.SUCCESS('Importación completada correctamente'))